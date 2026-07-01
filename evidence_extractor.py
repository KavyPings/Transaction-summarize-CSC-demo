from __future__ import annotations

import base64
import os
from pathlib import Path

EXCEL_EXTENSIONS = {".xlsx", ".xls"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".bmp", ".tiff"}

_MIME = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".bmp": "image/bmp",
    ".tiff": "image/tiff",
}


def is_image_file(path: str) -> bool:
    return Path(path).suffix.lower() in IMAGE_EXTENSIONS


def encode_image_for_llm(path: str) -> dict:
    """Return an OpenAI image_url content block for direct vision input."""
    suffix = Path(path).suffix.lower()
    mime = _MIME.get(suffix, "image/png")
    with open(path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode()
    return {
        "type": "image_url",
        "image_url": {"url": f"data:{mime};base64,{b64}", "detail": "high"},
    }


def _extract_excel(path: str) -> str:
    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise RuntimeError("Install openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True)
    sections: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
        if not rows:
            continue
        lines = [f"Sheet: {sheet_name}"]
        for row in rows:
            cells = [str(c) if c is not None else "" for c in row]
            while cells and not cells[-1]:
                cells.pop()
            if cells:
                lines.append("  |  ".join(cells))
        sections.append("\n".join(lines))

    return "\n\n".join(sections) if sections else "(empty workbook)"


def _extract_pdf(path: str) -> str:
    try:
        import pdfplumber  # type: ignore
    except ImportError:
        raise RuntimeError("Install pdfplumber: pip install pdfplumber")

    pages: list[str] = []
    with pdfplumber.open(path) as pdf:
        for i, page in enumerate(pdf.pages, 1):
            text = page.extract_text()
            if text and text.strip():
                pages.append(f"--- Page {i} ---\n{text.strip()}")

    return "\n\n".join(pages) if pages else "(no text found in PDF)"


def _extract_msg(path: str) -> str:
    try:
        import extract_msg  # type: ignore
    except ImportError:
        raise RuntimeError("Install extract-msg: pip install extract-msg")

    msg = extract_msg.Message(path)
    parts: list[str] = []
    if msg.subject:
        parts.append(f"Subject: {msg.subject}")
    if msg.sender:
        parts.append(f"From: {msg.sender}")
    if msg.date:
        parts.append(f"Date: {msg.date}")
    if msg.body:
        parts.append(f"\nBody:\n{msg.body.strip()}")
    return "\n".join(parts)


def build_user_message(transaction_json: str, evidence_paths: list[str]) -> object:
    """
    Builds the user message content for the LLM.
    Text evidences are embedded inline; images are sent as vision blocks.
    Returns a plain string when there are no images, or a content array otherwise.
    """
    if not evidence_paths:
        return transaction_json

    has_images = any(is_image_file(p) for p in evidence_paths)

    if not has_images:
        parts = [transaction_json]
        for i, path in enumerate(evidence_paths, 1):
            label = f"Evidence {i} ({Path(path).suffix.lstrip('.').upper()})"
            try:
                parts.append(f"\n\n{label}:\n{extract_evidence_text(path)}")
            except Exception as exc:
                parts.append(f"\n\n{label}: [Could not extract — {exc}]")
        return "".join(parts)

    content: list[dict] = [{"type": "text", "text": transaction_json}]
    for i, path in enumerate(evidence_paths, 1):
        label = f"Evidence {i} ({Path(path).suffix.lstrip('.').upper()})"
        if is_image_file(path):
            content.append({"type": "text", "text": f"\n\n{label}:"})
            content.append(encode_image_for_llm(path))
        else:
            try:
                text = extract_evidence_text(path)
                content.append({"type": "text", "text": f"\n\n{label}:\n{text}"})
            except Exception as exc:
                content.append({"type": "text", "text": f"\n\n{label}: [Could not extract — {exc}]"})
    return content


def extract_evidence_text(path: str) -> str:
    if not os.path.exists(path):
        raise FileNotFoundError(f"Evidence file not found: {path}")

    suffix = Path(path).suffix.lower()

    if suffix in EXCEL_EXTENSIONS:
        return _extract_excel(path)
    if suffix == ".pdf":
        return _extract_pdf(path)
    if suffix == ".msg":
        return _extract_msg(path)

    raise ValueError(f"Unsupported file type for text extraction: {suffix}")

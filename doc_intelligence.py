from __future__ import annotations

import os
from pathlib import Path

from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.core.credentials import AzureKeyCredential

CONTENT_TYPES: dict[str, str] = {
    ".pdf":  "application/pdf",
    ".png":  "image/png",
    ".jpg":  "image/jpeg",
    ".jpeg": "image/jpeg",
    ".bmp":  "image/bmp",
    ".tiff": "image/tiff",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xls":  "application/vnd.ms-excel",
}

EXCEL_EXTENSIONS = {".xlsx", ".xls"}



def _create_client() -> DocumentIntelligenceClient:
    endpoint = os.getenv("AZURE_DOC_INTELLIGENCE_ENDPOINT")
    key = os.getenv("AZURE_DOC_INTELLIGENCE_KEY")
    if not endpoint or not key:
        raise RuntimeError(
            "Missing AZURE_DOC_INTELLIGENCE_ENDPOINT or AZURE_DOC_INTELLIGENCE_KEY in .env"
        )
    return DocumentIntelligenceClient(endpoint=endpoint, credential=AzureKeyCredential(key))


def _format_table(table) -> str:
    grid: list[list[str]] = [
        [""] * table.column_count for _ in range(table.row_count)
    ]
    for cell in table.cells:
        grid[cell.row_index][cell.column_index] = cell.content.replace("\n", " ").strip()

    col_widths = [
        max(max((len(grid[r][c]) for r in range(table.row_count)), default=0), 8)
        for c in range(table.column_count)
    ]

    rows_out: list[str] = []
    for row_idx, row in enumerate(grid):
        row_str = "  |  ".join(cell.ljust(col_widths[ci]) for ci, cell in enumerate(row))
        rows_out.append(row_str)
        if row_idx == 0:
            rows_out.append("-" * len(row_str))
    return "\n".join(rows_out)




def _build_di_overview(result, filename: str) -> str:
    """
    Build a compact hierarchical overview from a prebuilt-layout result.
    Elements (paragraphs, tables, figures) are sorted by their span offset
    so the document order is preserved.
    """
    lines: list[str] = []

    page_count = len(result.pages) if result.pages else 0
    table_count = len(result.tables) if result.tables else 0
    figure_count = len(getattr(result, "figures", None) or [])

    lines.append(f"Document: {Path(filename).name}")
    lines.append(
        f"Pages: {page_count}  |  Tables: {table_count}  |  Figures: {figure_count}"
    )

    elements: list[dict] = []

    for para in result.paragraphs or []:
        offset = para.spans[0].offset if para.spans else 0
        elements.append({
            "type": "paragraph",
            "role": para.role or "body",
            "content": para.content,
            "offset": offset,
        })

    for i, table in enumerate(result.tables or []):
        offset = table.spans[0].offset if table.spans else 0
        elements.append({
            "type": "table",
            "idx": i + 1,
            "table": table,
            "offset": offset,
        })

    for i, figure in enumerate(getattr(result, "figures", None) or []):
        offset = figure.spans[0].offset if figure.spans else 0
        caption = (
            figure.caption.content
            if getattr(figure, "caption", None) and figure.caption
            else None
        )
        elements.append({
            "type": "figure",
            "idx": i + 1,
            "caption": caption,
            "offset": offset,
        })

    elements.sort(key=lambda x: x["offset"])

    for elem in elements:
        if elem["type"] == "paragraph":
            role = elem["role"]
            content = elem["content"].strip()
            if not content or role in ("pageHeader", "pageFooter", "pageNumber"):
                continue
            if role == "title":
                lines.append("")
                lines.append(f"# {content}")
            elif role == "sectionHeading":
                lines.append("")
                lines.append(f"## {content}")
            elif role == "footnote":
                lines.append(f"  * {content}")
            else:
                lines.append(content)

        elif elem["type"] == "table":
            t = elem["table"]
            lines.append("")
            lines.append(f"[TABLE {elem['idx']}  —  {t.row_count} rows x {t.column_count} columns]")
            lines.append(_format_table(t))

        elif elem["type"] == "figure":
            caption_str = f": {elem['caption']}" if elem["caption"] else ""
            lines.append(f"[FIGURE {elem['idx']}{caption_str}]")

    return "\n".join(lines)


def _process_with_di(path: str) -> str:
    suffix = Path(path).suffix.lower()
    content_type = CONTENT_TYPES.get(suffix)
    if not content_type:
        raise ValueError(f"Unsupported file type for Document Intelligence: {suffix}")

    client = _create_client()
    with open(path, "rb") as f:
        poller = client.begin_analyze_document(
            "prebuilt-layout",
            body=f,
            content_type=content_type,
        )
    return _build_di_overview(poller.result(), path)



def _process_excel_openpyxl(path: str) -> str:
    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise RuntimeError("Install openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True)
    sections: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw_rows = list(ws.iter_rows(values_only=True))

        
        data_rows = [r for r in raw_rows if any(c is not None for c in r)]
        if not data_rows:
            continue

       
        max_col = max(
            (i + 1 for row in data_rows for i, c in enumerate(row) if c is not None),
            default=0,
        )
        if max_col == 0:
            continue

        trimmed = [tuple(row[:max_col]) for row in data_rows]

        col_widths = [
            max(
                max(
                    (len(str(row[c])) for row in trimmed if c < len(row) and row[c] is not None),
                    default=0,
                ),
                8,
            )
            for c in range(max_col)
        ]

        sheet_lines: list[str] = [f"Sheet: {sheet_name}", "-" * 60]
        for row_idx, row in enumerate(trimmed):
            cells = [
                str(row[c]) if c < len(row) and row[c] is not None else ""
                for c in range(max_col)
            ]
            row_str = "  |  ".join(cell.ljust(col_widths[ci]) for ci, cell in enumerate(cells))
            sheet_lines.append(row_str)
            if row_idx == 0:
                sheet_lines.append("-" * len(row_str))

        sections.append("\n".join(sheet_lines))

    return "\n\n".join(sections) if sections else "(No data found in workbook)"



def _process_msg(path: str) -> str:
    try:
        import extract_msg  # type: ignore
    except ImportError:
        raise RuntimeError("Install extract-msg: pip install extract-msg")

    msg = extract_msg.Message(path)
    parts: list[str] = []
    if msg.subject:
        parts.append(f"Subject: {msg.subject}")
    if msg.sender:
        parts.append(f"From: {msg.sender}")
    if msg.date:
        parts.append(f"Date: {msg.date}")
    if msg.body:
        parts.append(f"\nBody:\n{msg.body.strip()}")
    return "\n".join(parts)



def get_excel_content(path: str) -> str:
    """Extract spreadsheet content via openpyxl (XLSX/XLS only)."""
    if not os.path.exists(path):
        raise FileNotFoundError(f"Evidence file not found: {path}")
    return _process_excel_openpyxl(path)


def get_document_overview(path: str) -> str:
    """Extract a hierarchical overview via Azure Document Intelligence."""
    if not os.path.exists(path):
        raise FileNotFoundError(f"Evidence file not found: {path}")

    suffix = Path(path).suffix.lower()
    if suffix == ".msg":
        return _process_msg(path)
    return _process_with_di(path)
